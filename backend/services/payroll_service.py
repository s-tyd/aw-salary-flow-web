"""
給与計算Excel生成サービス
Firebase Cloud Functionsからの移行版
"""
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from typing import List, Dict, Any, Optional
from decimal import Decimal
from io import BytesIO
import datetime
import logging

from models import (
    CalculationPeriod, Employee, AttendanceRecord, 
    FreeeExpense, KinconeTransportation, ExcelTemplate
)
from schemas import WorkDataSummary, PayrollGenerationResponse

logger = logging.getLogger(__name__)

class PayrollService:
    """給与計算Excel生成サービス"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def generate_payroll_excel(
        self, 
        calculation_period_id: int, 
        template_id: int
    ) -> PayrollGenerationResponse:
        """
        給与計算Excelファイルを生成
        Firebase Cloud Functionsのgenerate_payroll関数を移行
        """
        error_messages = []
        
        try:
            # 計算期間の取得
            calculation_period = self.db.query(CalculationPeriod).filter(
                CalculationPeriod.id == calculation_period_id
            ).first()
            
            if not calculation_period:
                error_messages.append('指定された計算期間が見つかりません。')
                return PayrollGenerationResponse(
                    status="error",
                    messages=error_messages
                )
            
            # Excelテンプレートの取得
            template = self.db.query(ExcelTemplate).filter(
                ExcelTemplate.id == template_id,
                ExcelTemplate.is_active == True
            ).first()
            
            if not template:
                error_messages.append('指定されたテンプレートが見つかりません。')
                return PayrollGenerationResponse(
                    status="error", 
                    messages=error_messages
                )
            
            # 従業員データの統合取得
            work_data_summaries = self._get_work_data_summaries(calculation_period_id)
            
            if not work_data_summaries:
                error_messages.append('指定された年月のデータが見つかりません。')
                return PayrollGenerationResponse(
                    status="error",
                    messages=error_messages
                )
            
            # テンプレートファイルの読み込み
            template_content = self._load_template_file(template)
            if not template_content:
                error_messages.append('テンプレートファイルの読み込みに失敗しました。')
                return PayrollGenerationResponse(
                    status="error",
                    messages=error_messages
                )
            
            # Excel処理
            template_df = pd.read_excel(template_content, header=4)
            template_wb = load_workbook(template_content)
            ws = template_wb.active
            
            # A列にある社員番号の列を取得
            employee_numbers = template_df.iloc[:, 0].astype(str)
            
            # 各従業員データをExcelに書き込み
            for work_data in work_data_summaries:
                employee_number_str = str(work_data.employee_number)
                
                # テンプレート内の該当行を検索
                matching_row_index = template_df[employee_numbers == employee_number_str].index
                if not matching_row_index.empty:
                    row = matching_row_index[0] + 5 + 1  # ヘッダー行を考慮した実際の行番号
                    
                    # 各種データの書き込み（元のFirebase実装と同じ位置）
                    self._write_work_data_to_excel(ws, row, work_data)
            
            # 生成されたExcelファイルをメモリに保存
            output_stream = BytesIO()
            template_wb.save(output_stream)
            output_stream.seek(0)
            
            # ファイル名生成
            dt_now = datetime.datetime.utcnow() + datetime.timedelta(hours=9)
            file_name = f'payroll_{calculation_period.year}_{calculation_period.month:02d}_{dt_now.strftime("%Y%m%d%H%M%S")}.xlsx'
            
            # ファイル保存処理
            file_path = self._save_generated_file(output_stream, file_name)
            download_url = f"/payroll/download/{file_name}" if file_path else None
            
            logger.info(f"給与計算Excel生成完了: {file_name}")
            if file_path:
                logger.info(f"ファイル保存先: {file_path}")
            
            return PayrollGenerationResponse(
                status="success",
                messages=error_messages,
                file_name=file_name,
                download_url=download_url
            )
            
        except Exception as e:
            logger.error(f'給与計算Excel生成エラー: {str(e)}')
            return PayrollGenerationResponse(
                status="error",
                messages=[str(e)]
            )
    
    def _get_work_data_summaries(self, calculation_period_id: int) -> List[WorkDataSummary]:
        """
        計算期間内の全従業員の勤務データを統合して取得
        """
        summaries = []
        
        # 従業員一覧の取得
        employees = self.db.query(Employee).filter(
            Employee.is_active == True
        ).all()
        
        for employee in employees:
            # 勤務時間データの取得
            attendance_record = self.db.query(AttendanceRecord).filter(
                AttendanceRecord.calculation_period_id == calculation_period_id,
                AttendanceRecord.employee_id == employee.id
            ).first()
            
            # Freee経費の合計
            freee_total = self.db.query(FreeeExpense).filter(
                FreeeExpense.calculation_period_id == calculation_period_id,
                FreeeExpense.employee_id == employee.id
            ).with_entities(
                FreeeExpense.amount
            ).all()
            freee_expenses = sum([expense.amount for expense in freee_total]) if freee_total else 0
            
            # Kincone交通費の合計
            kincone_total = self.db.query(KinconeTransportation).filter(
                KinconeTransportation.calculation_period_id == calculation_period_id,
                KinconeTransportation.employee_id == employee.id
            ).with_entities(
                KinconeTransportation.amount
            ).all()
            kincone_expenses = sum([transport.amount for transport in kincone_total]) if kincone_total else 0
            
            # WorkDataSummaryの作成
            summary = WorkDataSummary(
                employee_id=employee.id,
                employee_number=employee.employee_number,
                employee_name=employee.name,
                # 勤務時間データから（文字列のまま使用）
                working_days=attendance_record.work_days if attendance_record else None,
                total_work_hours=attendance_record.total_work_time if attendance_record and attendance_record.total_work_time else None,
                paid_leave_days=float(attendance_record.paid_leave_used) if attendance_record and attendance_record.paid_leave_used else None,
                statutory_holiday_hours=attendance_record.holiday_work_time if attendance_record and attendance_record.holiday_work_time else None,
                night_working_hours=attendance_record.late_night_work_time if attendance_record and attendance_record.late_night_work_time else None,
                absence_days=attendance_record.absence_days if attendance_record else None,
                # TODO: 以下のフィールドは将来の機能拡張で実装
                remote_count=None,
                lunch_count=None,
                office_count=None,
                event_count=None,
                trip_night_before_count=None,
                trip_count=None,
                travel_onday_count=None,
                travel_holidays_count=None,
                special_holiday=None,
                special_holiday_without_pay=None,
                kiwi_points=None,
                # 経費データから
                freee_expenses=int(freee_expenses),
                kincone_expenses=int(kincone_expenses),
                no_remote_allowance_limit=False  # TODO: 従業員マスタに追加予定
            )
            
            summaries.append(summary)
        
        return summaries
    
    def _save_generated_file(self, file_stream: BytesIO, file_name: str) -> Optional[str]:
        """
        生成されたExcelファイルをローカルに保存
        """
        try:
            import os
            
            # 出力ディレクトリの作成
            current_dir = os.path.dirname(os.path.dirname(__file__))  # backend/
            project_root = os.path.dirname(current_dir)  # プロジェクトルート
            output_dir = os.path.join(project_root, "output_files")
            
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
                logger.info(f"出力ディレクトリを作成: {output_dir}")
            
            # ファイル保存
            file_path = os.path.join(output_dir, file_name)
            file_stream.seek(0)  # ストリームの先頭に戻る
            
            with open(file_path, 'wb') as f:
                f.write(file_stream.read())
            
            logger.info(f"ファイル保存完了: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"ファイル保存エラー: {str(e)}")
            return None
    
    def _parse_time_to_timedelta(self, time_str: str) -> Optional[datetime.timedelta]:
        """
        時間文字列（HH:MM形式）をtimedelta型に変換
        """
        if not time_str or time_str == '-':
            return None
        
        try:
            if ':' in time_str:
                hours, minutes = time_str.split(':')
                return datetime.timedelta(hours=int(hours), minutes=int(minutes))
            else:
                return datetime.timedelta(hours=float(time_str))
        except ValueError:
            return None
    
    def _timedelta_to_hours(self, td) -> float:
        """
        timedelta型または文字列を時間（float）に変換
        """
        if not td:
            return 0.0
            
        logger.info(f"時間変換処理: {td} (型: {type(td)})")
            
        # 文字列の場合（PostgreSQLのINTERVAL型）
        if isinstance(td, str):
            hours = self._parse_interval_string_to_hours(td)
            logger.info(f"文字列から時間変換: '{td}' → {hours}時間")
            return hours
            
        # timedelta型の場合
        if isinstance(td, datetime.timedelta):
            hours = td.total_seconds() / 3600
            logger.info(f"timedeltaから時間変換: {td} → {hours}時間")
            return hours
            
        logger.warning(f"未対応の型での時間変換: {td} (型: {type(td)})")
        return 0.0
    
    def _parse_interval_string_to_hours(self, interval_str: str) -> float:
        """
        PostgreSQL INTERVAL文字列を時間（float）に変換
        例: '142:02:00' → 142.033333
        """
        if not interval_str:
            return 0.0
            
        logger.info(f"INTERVAL文字列解析開始: '{interval_str}'")
            
        try:
            # HH:MM:SS形式の場合
            if ':' in interval_str:
                parts = interval_str.split(':')
                hours = int(parts[0])
                minutes = int(parts[1]) if len(parts) > 1 else 0
                seconds = int(parts[2]) if len(parts) > 2 else 0
                
                result = hours + minutes / 60 + seconds / 3600
                logger.info(f"HH:MM:SS形式解析: {hours}h {minutes}m {seconds}s → {result}時間")
                return result
                
            # その他の形式（例: "142 hours 2 minutes"）
            import re
            
            # 時間を抽出
            hours_match = re.search(r'(\d+)\s*(?:hours?|hrs?|h)', interval_str, re.IGNORECASE)
            minutes_match = re.search(r'(\d+)\s*(?:minutes?|mins?|m)', interval_str, re.IGNORECASE)
            seconds_match = re.search(r'(\d+)\s*(?:seconds?|secs?|s)', interval_str, re.IGNORECASE)
            
            hours = int(hours_match.group(1)) if hours_match else 0
            minutes = int(minutes_match.group(1)) if minutes_match else 0
            seconds = int(seconds_match.group(1)) if seconds_match else 0
            
            result = hours + minutes / 60 + seconds / 3600
            logger.info(f"テキスト形式解析: {hours}h {minutes}m {seconds}s → {result}時間")
            return result
            
        except (ValueError, IndexError, AttributeError) as e:
            logger.warning(f"INTERVAL文字列の解析に失敗: '{interval_str}' - エラー: {str(e)}")
            return 0.0
    
    def _timedelta_to_minutes(self, td) -> int:
        """
        timedelta型または文字列を分（int）に変換
        """
        if not td:
            return 0
            
        # 文字列の場合（PostgreSQLのINTERVAL型）
        if isinstance(td, str):
            hours = self._parse_interval_string_to_hours(td)
            return int(hours * 60)
            
        # timedelta型の場合
        if isinstance(td, datetime.timedelta):
            return int(td.total_seconds() / 60)
            
        return 0
    
    def _format_timedelta_to_hm(self, td) -> str:
        """
        timedelta型または文字列を HH:MM 形式の文字列に変換
        """
        if not td:
            return "0:00"
            
        # 文字列の場合（PostgreSQLのINTERVAL型）
        if isinstance(td, str):
            # すでにHH:MM:SS形式の場合はそのまま使用
            if ':' in td:
                parts = td.split(':')
                hours = int(parts[0])
                minutes = int(parts[1]) if len(parts) > 1 else 0
                return f"{hours}:{minutes:02d}"
            else:
                # 他の形式の場合は時間に変換してからフォーマット
                hours = self._parse_interval_string_to_hours(td)
                total_minutes = int(hours * 60)
                h = total_minutes // 60
                m = total_minutes % 60
                return f"{h}:{m:02d}"
        
        # timedelta型の場合
        if isinstance(td, datetime.timedelta):
            total_minutes = int(td.total_seconds() / 60)
            hours = total_minutes // 60
            minutes = total_minutes % 60
            return f"{hours}:{minutes:02d}"
            
        return "0:00"
    
    def _format_hours_to_hm(self, hours: float) -> str:
        """
        時間（float）をHH:MM形式の文字列に変換
        """
        if not hours or hours <= 0:
            return "0:00"
            
        total_minutes = int(hours * 60)
        h = total_minutes // 60
        m = total_minutes % 60
        return f"{h}:{m:02d}"
    
    def _load_template_file(self, template: ExcelTemplate) -> Optional[BytesIO]:
        """
        Excelテンプレートファイルを読み込み
        データベースに保存されたfile_dataを使用
        """
        try:
            if template.file_data is None:
                logger.error(f"テンプレート ID {template.id} にファイルデータが保存されていません")
                return None
            
            logger.info(f"テンプレート読み込み: ID={template.id}, 名前={template.name}")
            logger.info(f"ファイルサイズ: {len(template.file_data)} bytes")
            
            # データベースからファイルデータを読み込み
            content = BytesIO(template.file_data)
            logger.info(f"テンプレートファイル読み込み成功: {len(template.file_data)} bytes")
            return content
                
        except Exception as e:
            logger.error(f"テンプレートファイル読み込みエラー: {str(e)}")
            return None
    
    def _write_work_data_to_excel(self, ws, row: int, work_data: WorkDataSummary):
        """
        勤務データをExcelの指定行に書き込み
        Firebase実装と同じ列位置を使用
        """
        # 出勤日数
        if work_data.working_days is not None:
            ws[f'E{row}'] = work_data.working_days
        
        # 総労働時間（文字列のまま書き込み）
        if work_data.total_work_hours is not None:
            ws[f'F{row}'] = work_data.total_work_hours
        
        # 有給取得日数
        if work_data.paid_leave_days is not None:
            ws[f'G{row}'] = work_data.paid_leave_days
        
        # 法定休日時間（文字列のまま書き込み）
        if work_data.statutory_holiday_hours is not None:
            ws[f'N{row}'] = work_data.statutory_holiday_hours
        
        # 深夜時間（文字列のまま書き込み）
        if work_data.night_working_hours is not None:
            ws[f'O{row}'] = work_data.night_working_hours
        
        # 欠勤日数
        if work_data.absence_days is not None:
            ws[f'P{row}'] = work_data.absence_days
        
        # リモート＠家（上限10日、noRemoteAllowanceLimitチェック）
        if work_data.remote_count is not None:
            remote_count = work_data.remote_count
            if remote_count >= 10 and not work_data.no_remote_allowance_limit:
                remote_count = 10
            ws[f'X{row}'] = remote_count
        
        # 出社ランチ（上限10日）
        if work_data.lunch_count is not None:
            lunch_count = work_data.lunch_count
            if lunch_count >= 10:
                lunch_count = 10
            ws[f'AG{row}'] = f'=500*{lunch_count}'
        
        # 通勤日数
        if work_data.office_count is not None:
            ws[f'AP{row}'] = work_data.office_count
        
        # オフィス出社手当（上限10日）
        if work_data.office_count is not None:
            office_count = work_data.office_count
            if office_count >= 10:
                office_count = 10
            ws[f'AF{row}'] = f'=2000*{office_count}'
        
        # イベント参加
        if work_data.event_count is not None:
            ws[f'AH{row}'] = f'=3000*{work_data.event_count}'
        
        # 出張日当計算
        if all(x is not None for x in [
            work_data.trip_night_before_count, work_data.trip_count,
            work_data.travel_onday_count, work_data.travel_holidays_count
        ]):
            ws[f'AR{row}'] = (
                f'=2000*{work_data.trip_night_before_count}+'
                f'2000*{work_data.trip_count}+'
                f'2000*{work_data.travel_onday_count}+'
                f'1000*{work_data.travel_holidays_count}'
            )
        
        # Kiwi points
        if work_data.kiwi_points is not None:
            ws[f'AK{row}'] = work_data.kiwi_points
        
        # 特別休暇
        if work_data.special_holiday is not None:
            ws[f'H{row}'] = f'=H2+{work_data.special_holiday}'
        
        # 特別休暇(無給)
        if work_data.special_holiday_without_pay is not None:
            ws[f'I{row}'] = f'=I2+{work_data.special_holiday_without_pay}'
        
        # freee立替経費
        if work_data.freee_expenses is not None:
            ws[f'AT{row}'] = work_data.freee_expenses
        
        # kincone立替経費
        if work_data.kincone_expenses is not None:
            ws[f'AU{row}'] = work_data.kincone_expenses