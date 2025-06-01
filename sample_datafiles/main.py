import firebase_admin
from firebase_admin import credentials, storage, db
from firebase_functions import https_fn
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import datetime
import logging

# ログ設定
logging.basicConfig(level=logging.DEBUG, filename='/tmp/app.log', filemode='a',
                    format='%(asctime)s - %(levelname)s - %(message)s')

cred = credentials.Certificate("salary-flow-agileware-firebase-adminsdk-zgw2v-9651806def.json")

# local
# firebase_admin.initialize_app(cred, {
#     'storageBucket': 'salary-flow-agileware.appspot.com',
#     'storage': {
#         'host': 'http://localhost:8080'
#     },
#     'databaseURL': 'http://127.0.0.1:9000/?ns=salary-flow-agileware-default-rtdb'
# })

firebase_admin.initialize_app(cred, {
    'storageBucket': 'salary-flow-agileware.appspot.com',
    'databaseURL': 'https://salary-flow-agileware-default-rtdb.firebaseio.com'
})


# # グローバルオプションの設定
# options.set_global_options(region=options.SupportedRegion.ASIA_NORTHEAST1)


def get_template_excel(date_path):
    # Firebase Storage からファイルを取得
    bucket = storage.bucket()
    blob = bucket.blob(f'{date_path}/template.xlsx')

    if blob.exists():
        # ファイルをバイトデータとしてダウンロード
        content = blob.download_as_bytes()
        excel_file = BytesIO(content)
        return excel_file


@https_fn.on_call()
def hello_world(req) -> https_fn.Response:
    return {"message": "Hello from Cloud Functions!", "data": req.data}


@https_fn.on_call()
def generate_payroll(req) -> https_fn.Response:
    error_messages = []

    try:
        data = req.data
        date_path = data['date_path']
        dt_now = datetime.datetime.utcnow() + datetime.timedelta(hours=9)

        # Realtime Database からデータを取得
        ref = db.reference(f'{date_path}/employees')
        employees_data = ref.get()
        if employees_data is None:
            error_messages.append('No data found for the specified year and month.')

        template_file = get_template_excel(date_path)
        if template_file is None:
            error_messages.append('No data found in the template file.')
        template_df = pd.read_excel(template_file, header=4)
        template_wb = load_workbook(template_file)
        ws = template_wb.active

        # A列にある社員番号の列を取得
        employee_numbers = template_df.iloc[:, 0].astype(str)  # 社員番号を文字列として扱う

        # DataFrameと照らし合わせて行に書き込み
        for key in employees_data.keys():
            try:
                key_str = str(int(key))
            except ValueError:
                key_str = str(key)

            employee = employees_data[key]

            if employee.get('approved', False) is False:
                continue

            work_data = employee.get('work_data', {})
            work_tag = work_data.get('workTag', {})

            matching_row_index = template_df[employee_numbers == key_str].index
            if not matching_row_index.empty:
                row = matching_row_index[0] + 5 + 1

                # 出勤日数
                working_days = work_data.get('workingDays')
                if working_days is not None:
                    ws[f'E{row}'] = working_days

                # 総労働時間
                total_work_hours = work_data.get('totalWorkingHours')
                if total_work_hours is not None:
                    ws[f'F{row}'] = total_work_hours

                # 有給取得日数
                paid_leave_days = work_data.get('paidLeaveDays')
                if paid_leave_days is not None:
                    ws[f'G{row}'] = paid_leave_days

                # 法定休日時間
                statutory_holiday_hours = work_data.get('statutoryHolidayHours')
                if statutory_holiday_hours is not None:
                    ws[f'N{row}'] = statutory_holiday_hours

                # 深夜時間
                night_working_hours = work_data.get('nightWorkingHours')
                if night_working_hours is not None:
                    ws[f'O{row}'] = night_working_hours

                # 欠勤日数
                absence_days = work_data.get('absence')
                if absence_days is not None:
                    ws[f'P{row}'] = absence_days

                # リモート＠家
                remote_count = int(work_tag.get('remote', 0))
                remote_nolimit = employee.get('name_keys', {}).get('noRemoteAllowanceLimit', False)
                if remote_count is not None:
                    if remote_count >= 10 and not remote_nolimit:
                        remote_count = 10
                    ws[f'X{row}'] = remote_count

                # 出社ランチ
                lunch_count = int(work_tag.get('lunch', 0))
                if lunch_count >= 10:
                    lunch_count = 10
                ws[f'AG{row}'] = f'=500*{lunch_count}'

                # 通勤日数
                office_count = int(work_tag.get('office', 0))
                ws[f'AP{row}'] = office_count

                # オフィス出社手当
                if office_count >= 10:
                    office_count = 10
                ws[f'AF{row}'] = f'=2000*{office_count}'

                # イベント参加
                event_count = int(work_tag.get('event', 0))
                ws[f'AH{row}'] = f'=3000*{event_count}'

                # 出張日当計算
                trip_night_before_count = int(work_tag.get('tripNightBefore', 0))
                trip_count = int(work_tag.get('trip', 0))
                travel_onday_count = int(work_tag.get('travelOnDay', 0))
                travel_holidays_count = int(work_tag.get('travelHolidays', 0))
                ws[
                    f'AR{row}'] = f'=2000*{trip_night_before_count}+2000*{trip_count}+2000*{travel_onday_count}+1000*{travel_holidays_count}'

                # Kiwi points
                total_points = int(employee.get('kiwi_score_report', {}).get('point', 0))
                ws[f'AK{row}'] = total_points

                # 特別休暇
                special_holiday = int(work_tag.get('specialHoliday', 0))
                ws[f'H{row}'] = f'=H2+{special_holiday}'

                # 特別休暇(無給)
                special_holiday_without_pay = int(work_tag.get('specialHolidayWithoutPay', 0))
                ws[f'I{row}'] = f'=I2+{special_holiday_without_pay}'

                # freee立替経費
                freee_amount = int(employee.get('freee_expenses', {}).get('expenses', 0))
                ws[f'AT{row}'] = freee_amount

                # kincone立替経費
                total_amount = int(employee.get('transportation_expenses', {}).get('expenses', 0))
                ws[f'AU{row}'] = total_amount

        # 生成されたExcelファイルをメモリに保存
        output_stream = BytesIO()
        template_wb.save(output_stream)
        output_stream.seek(0)

        # Firebase Storageにファイルをアップロード
        save_filename = f'output_{dt_now.strftime("%Y%m%d%H%M%S")}.xlsx'
        blob = storage.bucket().blob(f'{date_path}/{save_filename}')
        blob.upload_from_file(output_stream,
                              content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # ファイルのURLを取得
        file_url = blob.generate_signed_url(expiration=datetime.timedelta(hours=1), method='GET')

        return {"status": "success", "messages": error_messages, "file_url": file_url, "file_name": save_filename}
    except Exception as e:
        logging.error(f'Error in generate_payroll: {str(e)}')
        return {"status": "error", "message": str(e), "errors": error_messages}
